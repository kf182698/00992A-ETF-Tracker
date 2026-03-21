# etf_tracker.py — 下載 00992A 每日持股 → 清洗 → 抓當日收盤價(快取) →
# 雙軌保存（抓檔日 daily / 官方快照日 snapshots）+ 去重 + manifest 追蹤
import os, re, time, glob, json, shutil, hashlib, csv
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests

from playwright.sync_api import sync_playwright

from openpyxl import load_workbook
from pandas import ExcelWriter

# === 目錄 ===
DOWNLOAD_DIR = "downloads"
DATA_DIR     = "data"
SNAP_DATA_DIR= "data_snapshots"
SCREEN_DIR   = "screenshots"
PRICE_DIR    = "prices"
MANIFEST_DIR = "manifest"
DAILY_ARCHIVE_DIR = "archive/daily"
SNAPSHOT_DIR = "archive/snapshots"

for d in (DOWNLOAD_DIR, DATA_DIR, SNAP_DATA_DIR, SCREEN_DIR, PRICE_DIR, MANIFEST_DIR, DAILY_ARCHIVE_DIR, SNAPSHOT_DIR):
    Path(d).mkdir(parents=True, exist_ok=True)

# === 網址 ===
FUND_CODE = os.environ.get("FUND_CODE", "500")  # 00992A
ETF_URL   = os.environ.get("EZMONEY_URL", f"https://www.ezmoney.com.tw/ETF/Fund/Info?fundCode={FUND_CODE}")

# === 欄位別名（放寬） ===
ALIASES = {
    "code":   ["股票代號","證券代號","代號","代碼","股票代碼","證券代碼","Symbol","Ticker","Code","Stock Code"],
    "name":   ["股票名稱","證券名稱","名稱","Name","Stock Name","Security Name"],
    "shares": ["股數","持股股數","持有股數","Shares","Units","Quantity","張數"],
    "weight": ["持股權重","持股比例","權重","占比","比重(%)","占比(%)","Weight","Holding Weight","Portfolio Weight"],
    "close":  ["收盤價","收盤","價格","Price","Close","Closing Price"],
}

def _norm(s): return str(s).strip().replace("　","").replace("\u3000","")

def _download_excel():
    # Fetch from Capital Fund using Playwright and save as Excel file
    print("[etf_tracker] Fetching Capital API for 00992A...")
    ymd_compact = datetime.now().strftime("%Y%m%d")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(ETF_URL, wait_until="networkidle")
        js_code = """
        () => {
            return fetch("https://www.capitalfund.com.tw/CFWeb/api/etf/buyback", {
                method: "POST",
                headers: { "Content-Type": "application/json" },
                body: JSON.stringify({ fundId: "500", date: null })
            }).then(res => res.json());
        }
        """
        data = page.evaluate(js_code)
        browser.close()

    if not data or data.get("code") != 200 or not data.get("data") or not data["data"].get("stocks"):
        raise RuntimeError("取得 Capital JSON 失敗")

    stocks = data["data"]["stocks"]
    records = []
    for s in stocks:
        records.append({
            "股票代號": str(s["stocNo"]).strip(),
            "股票名稱": str(s["stocName"]).strip(),
            "股數": s["share"],
            "持股權重": s["weight"]
        })
    df = pd.DataFrame(records)
    
    # Save as temp excel to simulate the download
    out_path = os.path.join(DOWNLOAD_DIR, f"temp_{ymd_compact}.xlsx")
    df.to_excel(out_path, index=False)
    print(f"[etf_tracker] Emulated download saved to {out_path}")
    return out_path


def _find_header_row(df):
    best_idx, best = None, {}
    for ridx in range(min(50,len(df))):
        row = df.iloc[ridx]
        m={}
        for cidx,val in enumerate(row):
            lab=_norm(val)
            if not lab or lab.startswith("Unnamed"): continue
            low=lab.lower()
            def hit(keys): return any(k.lower() in low for k in keys)
            if hit(ALIASES["code"])   and "code" not in m:   m["code"]=cidx
            if hit(ALIASES["name"])   and "name" not in m:   m["name"]=cidx
            if hit(ALIASES["shares"]) and "shares" not in m: m["shares"]=cidx
            if hit(ALIASES["weight"]) and "weight" not in m: m["weight"]=cidx
            if hit(ALIASES["close"])  and "close" not in m:  m["close"]=cidx
        score = sum(k in m for k in ("code","name","weight")) + (1 if "shares" in m else 0)
        if score>=2 and (best_idx is None or len(m)>len(best)):
            best_idx, best = ridx, m
    return best_idx, best

def _extract_table(xlsx_path):
    df0 = pd.read_excel(xlsx_path)
    df0.columns = [_norm(c) for c in df0.columns]
    def map_cols(cols):
        m={}
        for i,col in enumerate(cols):
            low=str(col).lower()
            if any(k.lower() in low for k in ALIASES["code"])   and "code" not in m:   m["code"]=i
            if any(k.lower() in low for k in ALIASES["name"])   and "name" not in m:   m["name"]=i
            if any(k.lower() in low for k in ALIASES["shares"]) and "shares" not in m: m["shares"]=i
            if any(k.lower() in low for k in ALIASES["weight"]) and "weight" not in m: m["weight"]=i
            if any(k.lower() in low for k in ALIASES["close"])  and "close" not in m:  m["close"]=i
        return m
    mapped = map_cols(df0.columns)
    if sum(k in mapped for k in ("code","name","weight"))<2:
        df1 = pd.read_excel(xlsx_path, header=None).applymap(_norm)
        idx, m2 = _find_header_row(df1)
        if idx is None: raise ValueError("無法辨識表頭")
        cols = df1.iloc[idx].tolist()
        body = df1.iloc[idx+1:].reset_index(drop=True)
        body.columns=[_norm(c) for c in cols]
        df0 = body
        mapped = map_cols(df0.columns)
    # 合欄拆解
    if "code" not in mapped and "name" in mapped:
        name_col = df0.columns[mapped["name"]]
        s = df0[name_col].astype(str)
        a = s.str.extract(r"^\s*(\d{4,6})\s*([^\d].*)$")
        b = s.str.extract(r"^(.+?)\s*[\(（](\d{4,6})[\)）]\s*$")
        if a.notna().all(1).sum() >= b.notna().all(1).sum():
            df0["_code"]=a[0]; df0["_name"]=a[1]
        else:
            df0["_code"]=b[1]; df0["_name"]=b[0]
        mapped["code"]=df0.columns.get_loc("_code"); mapped["name"]=df0.columns.get_loc("_name")

    need=[]
    for k in ("code","name","shares","weight"):
        if k in mapped: need.append(df0.columns[mapped[k]])
    if "code" not in mapped or "name" not in mapped or "weight" not in mapped:
        raise ValueError(f"欄位不足，columns={list(df0.columns)[:10]} mapped={mapped}")

    df = df0[need].copy()
    # 正式欄名
    new=[]
    for c in df.columns:
        low=str(c).lower()
        if any(k.lower() in low for k in ALIASES["code"]):   new.append("股票代號")
        elif any(k.lower() in low for k in ALIASES["name"]): new.append("股票名稱")
        elif any(k.lower() in low for k in ALIASES["shares"]): new.append("股數")
        elif any(k.lower() in low for k in ALIASES["weight"]): new.append("持股權重")
        else: new.append(c)
    df.columns=new

    df["股票代號"]=df["股票代號"].astype(str).str.strip()
    df["股票名稱"]=df["股票名稱"].astype(str).str.strip()
    df["股數"]=pd.to_numeric(df.get("股數",0).astype(str).str.replace(",","",regex=False),errors="coerce").fillna(0).astype(int)
    df["持股權重"]=pd.to_numeric(df["持股權重"].astype(str).str.replace(",","",regex=False).str.replace("%","",regex=False),errors="coerce").fillna(0.0)
    df = df[(df["股票代號"].str.match(r"^\d{4,6}$")) & (df["股票名稱"].str.len()>0)].reset_index(drop=True)
    return df

# === Yahoo 價格抓取與快取(json) + 當日 CSV 價格表 ===
def _yahoo_quote(codes):
    out={}
    sess = requests.Session()
    headers={"User-Agent":"Mozilla/5.0"}
    for code in codes:
        syms = [f"{code}.TW", f"{code}.TWO"]
        price=None
        for s in syms:
            url = "https://query1.finance.yahoo.com/v7/finance/quote"
            try:
                r = sess.get(url, params={"symbols": s}, timeout=10, headers=headers)
                if r.status_code!=200: continue
                js=r.json()
                res = js.get("quoteResponse",{}).get("result",[])
                if not res: continue
                p = res[0].get("regularMarketPrice") or res[0].get("postMarketPrice")
                if p: price=float(p); break
            except Exception: continue
        if price is not None: out[code]=price
    return out

def _load_price_cache(ymd):
    p = os.path.join(PRICE_DIR, f"{ymd}.json")
    if os.path.exists(p):
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: return {}
    return {}

def _save_price_cache(ymd, data):
    p = os.path.join(PRICE_DIR, f"{ymd}.json")
    with open(p,"w",encoding="utf-8") as f: json.dump(data,f,ensure_ascii=False,indent=2)

def _fetch_prices_for(df, ymd):
    cache = _load_price_cache(ymd)
    need = sorted({c for c in df["股票代號"].astype(str) if str(c) not in cache})
    if need:
        got = _yahoo_quote(need)
        cache.update(got)
        _save_price_cache(ymd, cache)
    closes=[]
    for code in df["股票代號"].astype(str):
        if code in cache:
            closes.append(cache[code])
        else:
            prev_files = sorted(glob.glob(os.path.join(PRICE_DIR,"*.json")))
            prev_files = [p for p in prev_files if os.path.basename(p).split(".")[0] < ymd]
            prev_files.sort(reverse=True)
            val=None
            for pf in prev_files:
                try:
                    js=json.load(open(pf,"r",encoding="utf-8"))
                    if str(code) in js: val=js[str(code)]; break
                except: pass
            closes.append(val if val is not None else None)
    return closes

def _save_price_csv(date_str, df):
    Path(PRICE_DIR).mkdir(parents=True, exist_ok=True)
    out = os.path.join(PRICE_DIR, f"{date_str}.csv")
    px = df[["股票代號", "收盤價"]].copy()
    px["股票代號"] = px["股票代號"].astype(str).str.strip()
    px.to_csv(out, index=False, encoding="utf-8-sig")
    print("[etf_tracker] saved prices csv:", out)

def _append_prices_sheet(xlsx_path, df):
    try:
        wb = load_workbook(xlsx_path)
        if "with_prices" in wb.sheetnames:
            wb.remove(wb["with_prices"]); wb.save(xlsx_path)
        with ExcelWriter(xlsx_path, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, sheet_name="with_prices", index=False)
        print("[etf_tracker] wrote sheet 'with_prices' into:", xlsx_path)
    except Exception as e:
        print("[etf_tracker] write with_prices failed:", e)

# === 快照日推斷 + 去重 ===
def _infer_snapshot_date_from_name(path_or_name: str, fallback_ymd: str) -> str:
    b = os.path.basename(path_or_name)
    m = re.search(r"(\d{8})", b)
    if not m: 
        return fallback_ymd
    y, mo, d = m.group(1)[0:4], m.group(1)[4:6], m.group(1)[6:8]
    return f"{y}-{mo}-{d}"

def _hash_df(df: pd.DataFrame) -> str:
    key_cols = [c for c in ("股票代號","股數","持股權重") if c in df.columns]
    arr = df[key_cols].copy()
    arr["股票代號"] = arr["股票代號"].astype(str).str.strip()
    arr["股數"] = pd.to_numeric(arr.get("股數",0), errors="coerce").fillna(0).astype(int)
    arr["持股權重"] = pd.to_numeric(arr.get("持股權重",0.0), errors="coerce").fillna(0.0).round(6)
    arr = arr.sort_values("股票代號").reset_index(drop=True)
    raw = arr.to_csv(index=False).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()

def _last_snapshot_hash():
    mf = os.path.join(MANIFEST_DIR, "snapshots.csv")
    if not os.path.exists(mf): return None
    try:
        last = None
        import csv as _csv
        with open(mf, newline="", encoding="utf-8") as f:
            for row in _csv.DictReader(f):
                last = row
        return (last or {}).get("hash")
    except Exception:
        return None

def _append_manifest(record: dict):
    mf = os.path.join(MANIFEST_DIR, "snapshots.csv")
    header = ["fetch_date","snapshot_date","rows","weight_sum","hash","daily_path","snapshot_path","is_new_snapshot"]
    write_header = not os.path.exists(mf)
    import csv as _csv
    with open(mf, "a", newline="", encoding="utf-8") as f:
        w = _csv.DictWriter(f, fieldnames=header)
        if write_header: w.writeheader()
        w.writerow({k: record.get(k, "") for k in header})

def main():
    ymd = datetime.now().strftime("%Y-%m-%d")
    raw_download = _download_excel()  # 原始下載檔名（含官方日期）
    snapshot_date = _infer_snapshot_date_from_name(raw_download, ymd)

    # 將下載檔整理成固定名字（downloads/YYYY-MM-DD.xlsx）
    fixed = os.path.join(DOWNLOAD_DIR, f"{ymd}.xlsx")
    try:
        if os.path.exists(fixed): os.remove(fixed)
        shutil.move(raw_download, fixed)
    except Exception as e:
        print("[etf_tracker] rename failed:", e); fixed = raw_download
    print("[etf_tracker] saved excel:", fixed)

    # 同步保存一份以抓檔日命名的 daily 原始檔
    daily_xlsx = os.path.join(DAILY_ARCHIVE_DIR, f"{ymd}.xlsx")
    try:
        if os.path.exists(daily_xlsx): os.remove(daily_xlsx)
        shutil.copy2(fixed, daily_xlsx)
    except Exception as e:
        print("[etf_tracker] copy daily xlsx failed:", e)

    # 解析表格
    df = _extract_table(fixed)

    # 抓價 + 寫入 df
    ymd_compact = ymd.replace("-", "")
    closes = _fetch_prices_for(df, ymd_compact)
    df["收盤價"] = pd.to_numeric(pd.Series(closes), errors="coerce")

    # 寫入每日 CSV（附來源快照日，供日後去重/分析）
    df_with_src = df.copy()
    df_with_src["source_snapshot_date"] = snapshot_date
    csv_out = os.path.join(DATA_DIR, f"{ymd}.csv")
    df_with_src.to_csv(csv_out, index=False, encoding="utf-8-sig")
    _save_price_csv(ymd, df_with_src)

    # 把 with_prices 工作表寫回固定下載檔與 daily 檔
    _append_prices_sheet(fixed, df_with_src)
    try:
        _append_prices_sheet(daily_xlsx, df_with_src)
    except Exception as e:
        print("[etf_tracker] write with_prices(daily copy) failed:", e)

    # 判斷是否新快照（用內容 hash 去重）
    h = _hash_df(df)
    last_h = _last_snapshot_hash()
    is_new = (h != last_h)

    snapshot_path = ""
    if is_new:
        # 以「快照日」命名保存一份 canonical
        snap_name = f"ETF_Investment_Portfolio_{snapshot_date.replace('-','')}.xlsx"
        snapshot_path = os.path.join(SNAPSHOT_DIR, snap_name)
        try:
            shutil.copy2(fixed, snapshot_path)
            _append_prices_sheet(snapshot_path, df_with_src)
        except Exception as e:
            print("[etf_tracker] save snapshot xlsx failed:", e)
        # 也輸出以快照日命名的一份 CSV（供分析直接使用）
        snap_csv = os.path.join(SNAP_DATA_DIR, f"{snapshot_date}.csv")
        df.to_csv(snap_csv, index=False, encoding="utf-8-sig")

    # 記錄 manifest
    record = {
        "fetch_date": ymd,
        "snapshot_date": snapshot_date,
        "rows": len(df),
        "weight_sum": round(float(df["持股權重"].sum()), 6),
        "hash": h,
        "daily_path": daily_xlsx,
        "snapshot_path": snapshot_path,
        "is_new_snapshot": "1" if is_new else "0",
    }
    _append_manifest(record)

    print("[etf_tracker] saved daily:", csv_out, "| snapshot_new:", is_new, snapshot_date)

if __name__ == "__main__":
    main()

